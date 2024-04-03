import os
from openpyxl import load_workbook
from time import sleep
from dotenv import load_dotenv
from selenium import webdriver
import re
from selenium.webdriver.common.by import By
from selenium.webdriver.firefox.service import Service as FirefoxService
from selenium.webdriver.firefox.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.firefox import GeckoDriverManager

load_dotenv()
GH_TOKEN = os.getenv("GH_TOKEN")
WEAPONS_CATEGORIES_URL = "https://darksouls.fandom.com/wiki/Weapons_(Dark_Souls)"
EXCEL_PATH = "SoulsData.xlsx"
BASE_WEAPON_PATH = "https://darksouls.fandom.com/wiki/"

def get_weapon_categories(driver):
    driver.get(WEAPONS_CATEGORIES_URL)
    span_elements = driver.find_elements(By.CLASS_NAME, "mw-headline")
    span_elements = span_elements[3:-2]
    wb = load_workbook(EXCEL_PATH)
    ws = wb["CategoriasArmas"]
    next_row = ws.max_row + 1
    for span in span_elements:
        match = re.search(r'<a[^>]*>(.*?)</a>', span.get_attribute("outerHTML"))
        if match:
            text = match.group(1)
            ws.cell(row=next_row, column=2, value=text)
            next_row += 1
    wb.save(EXCEL_PATH)

def get_weapon_images(driver):
    driver.get(WEAPONS_CATEGORIES_URL)
    div_elements = driver.find_elements(By.CLASS_NAME, "thumb")
    wb = load_workbook(EXCEL_PATH)
    ws = wb["Armas"]
    row_index = 2
    for div_element in div_elements:
        img = div_element.find_element(By.TAG_NAME, 'img')
        src = img.get_attribute('data-src')
        ws.cell(row=row_index, column=2).value = src
        row_index += 1
    wb.save(EXCEL_PATH)
    

def get_weapon_names(driver):
    driver.get(WEAPONS_CATEGORIES_URL)
    div_elements = driver.find_elements(By.CLASS_NAME, "lightbox-caption")
    workbook = load_workbook(EXCEL_PATH)
    sheet = workbook["Armas"]
    row_index = 2
    for div_element in div_elements:
        html_content = div_element.get_attribute("outerHTML")
        print(html_content)
        pattern = r'<a.*?>(.*?)<\/a>'
        matches = re.findall(pattern, html_content)
        extracted_text = matches[0] if matches else None
        print(extracted_text)
        
        sheet.cell(row=row_index, column=1).value = extracted_text
        row_index += 1
    workbook.save(EXCEL_PATH)

def get_weapons_physical_damage(driver):
    driver.get(WEAPONS_CATEGORIES_URL)
    physical_damage = []
    weapon_urls = []
    div_elements = driver.find_elements(By.CLASS_NAME, "lightbox-caption")
    for div_element in div_elements:
        a_tag_html = div_element.find_element(By.TAG_NAME, 'a').get_attribute("outerHTML")
        text_content = re.search(r'>(.*?)<\/a>', a_tag_html).group(1).replace(" ", "_")
        weapon_urls.append(BASE_WEAPON_PATH + text_content)
        
    for url in weapon_urls:
        try:
            driver.get(url)
            css_selectors = ['.mw-parser-output > ul:nth-child(3) > li:nth-child(1) > ul:nth-child(2) > li:nth-child(1) > a:nth-child(1)',
                             '.mw-parser-output > ul:nth-child(3) > li:nth-child(1) > a:nth-child(1)']
            for css_selector in css_selectors:
                try:
                    driver.find_element(By.CSS_SELECTOR, css_selector)
                    url += "_(Dark_Souls)"
                    driver.get(url)
                    break
                except:
                    continue

            damage_element = driver.find_element(By.CSS_SELECTOR, 'td.pi-horizontal-group-item.pi-data-value.pi-font.pi-border-color.pi-item-spacing[data-source="atk-physical"]')
            damage = damage_element.text
            print(url + " -> " + damage)
            physical_damage.append(damage if damage else "-")
        except:
            print("Element not found on the page.")
            physical_damage.append("-")

    for damage in physical_damage:
        print(damage)
        
  
def ds1_scrape():
    driver = webdriver.Firefox(
        service=FirefoxService(GeckoDriverManager().install())
    )
    #get_weapon_categories(driver)
    #get_weapon_images(driver)
    #get_weapon_names(driver)
    get_weapons_physical_damage(driver)
    driver.quit()


