import os
from openpyxl import load_workbook
from time import sleep
from dotenv import load_dotenv
from selenium import webdriver
import re
from selenium.webdriver.common.by import By
from selenium.webdriver.firefox.service import Service as FirefoxService
from selenium.webdriver.firefox.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.firefox import GeckoDriverManager

load_dotenv()
GH_TOKEN = os.getenv("GH_TOKEN")
WEAPONS_CATEGORIES_URL = "https://darksouls.fandom.com/wiki/Weapons_(Dark_Souls_II)"
EXCEL_PATH = "SoulsData.xlsx"
BASE_WEAPON_PATH = "https://darksouls.fandom.com/wiki/"
WEAPON_URLS = []

def get_weapon_urls_from_wiki(driver):
    driver.get(WEAPONS_CATEGORIES_URL)  
    sleep(2)
    driver.find_element(By.CSS_SELECTOR, "._2O--J403t2VqCuF8XJAZLK").click() 
    table1 = driver.find_element(By.XPATH, "/html/body/div[4]/div[4]/div[3]/main/div[3]/div/div/table[1]/tbody/tr")
    li_elements1 = table1.find_elements(By.TAG_NAME, "li")
    for li_element in li_elements1:
        a_tag = li_element.find_element(By.TAG_NAME, "a")
        url_suffix = a_tag.get_attribute("href")
        if "%27" in url_suffix:
            url_suffix = url_suffix.replace("%27", "'")
        WEAPON_URLS.append(url_suffix)
            
    table2 = driver.find_element(By.XPATH, "/html/body/div[4]/div[4]/div[3]/main/div[3]/div/div/table[2]/tbody/tr")
    li_elements2 = table2.find_elements(By.TAG_NAME, "li")
    for li_element in li_elements2:
        a_tag = li_element.find_element(By.TAG_NAME, "a")
        url_suffix = a_tag.get_attribute("href")
        if "%27" in url_suffix:
            url_suffix = url_suffix.replace("%27", "'")
        WEAPON_URLS.append(url_suffix)
    
    return WEAPON_URLS

def get_staff_urls_from_wiki(driver):
    driver.get(WEAPONS_CATEGORIES_URL)
    sleep(2)
    staves = driver.find_element(By.XPATH, "/html/body/div[4]/div[4]/div[3]/main/div[3]/div/div/table[2]/tbody/tr/td[1]")
    staves_li = staves.find_elements(By.TAG_NAME, "li")
    for staff in staves_li:
        a_tag = staff.find_element(By.TAG_NAME, "a")
        url_suffix = a_tag.get_attribute("href") 
        if "%27" in url_suffix:
            url_suffix = url_suffix.replace("%27", "'")
        WEAPON_URLS.append(url_suffix)
    return WEAPON_URLS

def get_chimes_urls_from_wiki(driver):
    driver.get(WEAPONS_CATEGORIES_URL)
    sleep(2)
    chimes = driver.find_element(By.XPATH, "/html/body/div[4]/div[4]/div[3]/main/div[3]/div/div/table[2]/tbody/tr/td[2]")
    chimes_li = chimes.find_elements(By.TAG_NAME, "li")
    for chime in chimes_li:
        a_tag = chime.find_element(By.TAG_NAME, "a")
        url_suffix = a_tag.get_attribute("href") 
        if "%27" in url_suffix:
            url_suffix = url_suffix.replace("%27", "'")
        WEAPON_URLS.append(url_suffix)
    return WEAPON_URLS

def get_weapon_categories(driver):
    driver.get(WEAPONS_CATEGORIES_URL)   
    span_elements = driver.find_elements(By.CLASS_NAME, "mw-headline")
    wb = load_workbook(EXCEL_PATH)
    ws = wb["CategoriasArmas"]
    start_row = 24
    for index, span in enumerate(span_elements, start=start_row):
        match = re.search(r'<a[^>]*>(.*?)</a>', span.get_attribute("outerHTML"))
        if match:
            text = match.group(1)
            ws.cell(row=index, column=2, value=text)
            print(text)
    wb.save(EXCEL_PATH)

def get_weapon_names(driver):
    driver.get(WEAPONS_CATEGORIES_URL)   
    li_elements = driver.find_elements(By.TAG_NAME, "li")[52:-8]
    for li_element in li_elements:
        match = re.search(r'<a[^>]*>(.*?)</a>', li_element.get_attribute("outerHTML"))
        if match: 
            text = match.group(1)
            print(text)

def get_weapon_images(driver):
    get_weapon_urls_from_wiki(driver)
    wb = load_workbook(EXCEL_PATH)
    ws = wb["Armas"]
    row = 149
    for weapon_url in WEAPON_URLS:
        driver.get(weapon_url)
        img_element = driver.find_element(By.CLASS_NAME, "pi-image-thumbnail")
        img_src = img_element.get_attribute("src")
        ws.cell(row=row, column=2).value = img_src
        row += 1
    wb.save(EXCEL_PATH)
    print("done")

def get_weapon_damages(driver):
    get_weapon_urls_from_wiki(driver)
    wb = load_workbook(EXCEL_PATH)
    ws = wb["Armas"]
    row = 149
    for weapon_url in WEAPON_URLS:
        driver.get(weapon_url)
        physical_damage = driver.find_element(By.CSS_SELECTOR, "section.pi-item:nth-child(5) > table:nth-child(1) > tbody:nth-child(2) > tr:nth-child(1) > td:nth-child(1)")
        match = re.search(r'<td[^>]*>(.*?)</td>', physical_damage.get_attribute("outerHTML"))
        if match: 
            damage = match.group(1)
            ws.cell(row=row, column=3).value = damage
            row += 1
    wb.save(EXCEL_PATH)
    print("done") 
    
def get_staff_damage(driver):
    get_staff_urls_from_wiki(driver)
    wb = load_workbook(EXCEL_PATH)
    ws = wb["Armas"]
    row = 371
    for staff_url in WEAPON_URLS:
        driver.get(staff_url)
        magical_damage = driver.find_element(By.CSS_SELECTOR, "section.pi-item:nth-child(5) > table:nth-child(1) > tbody:nth-child(2) > tr:nth-child(1) > td:nth-child(2)")
        match = re.search(r'<td[^>]*>(.*?)</td>', magical_damage.get_attribute("outerHTML"))
        if match: 
            damage = match.group(1)
            ws.cell(row=row, column=3).value = damage
            row += 1
    wb.save(EXCEL_PATH)
    print("done") 
    
def get_chime_damage(driver):
    get_chimes_urls_from_wiki(driver)
    wb = load_workbook(EXCEL_PATH)
    ws = wb["Armas"]
    row = 385
    for chime_url in WEAPON_URLS:
        driver.get(chime_url)
        dark_damage = driver.find_element(By.CSS_SELECTOR, "section.pi-item:nth-child(5) > table:nth-child(1) > tbody:nth-child(2) > tr:nth-child(1) > td:nth-child(5)")
        match = re.search(r'<td[^>]*>(.*?)</td>', dark_damage.get_attribute("outerHTML"))
        if match: 
            damage = match.group(1)
            ws.cell(row=row, column=3).value = damage
            row += 1
    wb.save(EXCEL_PATH)
    print("done") 

def ds2_scrape():
    driver = webdriver.Firefox(
        service=FirefoxService(GeckoDriverManager().install())
    )
    #get_weapon_categories(driver)
    #get_weapon_names(driver)
    #get_weapon_images(driver)
    #get_weapon_urls_from_wiki(driver)
    #get_weapon_damages(driver)
    #get_staff_damage(driver)
    get_chime_damage(driver)
    driver.quit()